import os
import customtkinter as ctk
from openpyxl import load_workbook
from config import COLORS, FONT_FAMILY, ATTENDANCE_FOLDER


class AttendanceViewerScreen(ctk.CTkFrame):
    def __init__(self, parent, app):
        super().__init__(parent, fg_color="transparent")
        self._app = app
        self._files = {}
        self._current_file = None

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # ── Header ──
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.grid(row=0, column=0, padx=20, pady=(16, 0), sticky="ew")
        header.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            header, text="Attendance Records",
            font=ctk.CTkFont(family=FONT_FAMILY, size=24, weight="bold"),
            text_color=COLORS["text_primary"],
        ).grid(row=0, column=0, sticky="w")

        # Controls
        controls = ctk.CTkFrame(header, fg_color="transparent")
        controls.grid(row=0, column=2, sticky="e")

        self._file_selector = ctk.CTkOptionMenu(
            controls, values=["Current Session"],
            font=ctk.CTkFont(family=FONT_FAMILY, size=12), width=260, height=34,
            fg_color=COLORS["bg_card"], button_color=COLORS["bg_elevated"],
            button_hover_color=COLORS["primary"],
            dropdown_fg_color=COLORS["bg_card"],
            dropdown_hover_color=COLORS["bg_elevated"],
            corner_radius=10,
            command=self._on_file_selected,
        )
        self._file_selector.grid(row=0, column=0, padx=4)

        ctk.CTkButton(
            controls, text="Refresh", width=80, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=COLORS["bg_elevated"], hover_color=COLORS["primary"],
            corner_radius=10, command=self._refresh,
        ).grid(row=0, column=1, padx=4)

        self._open_btn = ctk.CTkButton(
            controls, text="Open in Excel", width=110, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=COLORS["success"], hover_color=COLORS["success_hover"],
            corner_radius=10, command=self._open_file,
        )
        self._open_btn.grid(row=0, column=2, padx=4)

        ctk.CTkButton(
            controls, text="Open Folder", width=100, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=COLORS["bg_card"], hover_color=COLORS["bg_elevated"],
            border_width=1, border_color=COLORS["border"],
            corner_radius=10, command=self._open_folder,
        ).grid(row=0, column=3, padx=4)

        # ── Summary Bar ──
        summary = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=12, height=50,
                                border_width=1, border_color=COLORS["border"])
        summary.grid(row=1, column=0, padx=20, pady=(12, 0), sticky="ew")
        summary.grid_propagate(False)
        summary.grid_columnconfigure(3, weight=1)

        self._total_label = ctk.CTkLabel(
            summary, text="Total: 0 records",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            text_color=COLORS["text_primary"],
        )
        self._total_label.grid(row=0, column=0, padx=16, pady=12, sticky="w")

        self._file_info = ctk.CTkLabel(
            summary, text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=COLORS["text_muted"],
        )
        self._file_info.grid(row=0, column=3, padx=16, pady=12, sticky="e")

        # ── Table ──
        table_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=14,
                                    border_width=1, border_color=COLORS["border"])
        table_frame.grid(row=2, column=0, padx=20, pady=(10, 20), sticky="nsew")
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(1, weight=1)

        # Table Header
        header_row = ctk.CTkFrame(table_frame, fg_color=COLORS["bg_table_header"],
                                   corner_radius=0, height=42)
        header_row.grid(row=0, column=0, sticky="ew", padx=2, pady=(2, 0))
        header_row.grid_propagate(False)

        cols = ["#", "Student Name", "Date", "Time"]
        weights = [1, 5, 3, 2]
        for i, (col, w) in enumerate(zip(cols, weights)):
            header_row.grid_columnconfigure(i, weight=w)
            ctk.CTkLabel(
                header_row, text=col,
                font=ctk.CTkFont(family=FONT_FAMILY, size=11, weight="bold"),
                text_color=COLORS["text_muted"],
            ).grid(row=0, column=i, padx=14, pady=10, sticky="w")

        # Table Body
        self._table_body = ctk.CTkScrollableFrame(
            table_frame, fg_color="transparent",
            scrollbar_button_color=COLORS["bg_elevated"],
            scrollbar_button_hover_color=COLORS["primary"],
        )
        self._table_body.grid(row=1, column=0, sticky="nsew", padx=2, pady=(0, 2))
        for i, w in enumerate(weights):
            self._table_body.grid_columnconfigure(i, weight=w)

    def on_show(self):
        self._refresh()

    def on_hide(self):
        pass

    def _refresh(self):
        self._files = {"Current Session": None}
        if os.path.exists(ATTENDANCE_FOLDER):
            for f in sorted(os.listdir(ATTENDANCE_FOLDER), reverse=True):
                if f.endswith(".xlsx"):
                    display = f.replace("attendance_", "").replace(".xlsx", "").replace("_", "  ")
                    self._files[display] = os.path.join(ATTENDANCE_FOLDER, f)

        self._file_selector.configure(values=list(self._files.keys()))
        self._on_file_selected(self._file_selector.get())

    def _on_file_selected(self, choice):
        self._current_file = self._files.get(choice)
        self._load_records(choice)

    def _load_records(self, choice):
        for w in self._table_body.winfo_children():
            w.destroy()

        records = []
        if choice == "Current Session":
            for r in self._app.attendance_session.get_records():
                records.append(r)
            self._file_info.configure(text="Live session data")
        elif self._current_file and os.path.exists(self._current_file):
            records = self._read_excel(self._current_file)
            self._file_info.configure(text=os.path.basename(self._current_file))
        else:
            self._file_info.configure(text="")

        self._total_label.configure(text=f"Total: {len(records)} records")

        if not records:
            ctk.CTkLabel(
                self._table_body, text="No attendance records found",
                font=ctk.CTkFont(family=FONT_FAMILY, size=14),
                text_color=COLORS["text_muted"],
            ).grid(row=0, column=0, columnspan=4, pady=50)
            return

        for idx, r in enumerate(records):
            bg = COLORS["bg_table_stripe"] if idx % 2 == 0 else "transparent"
            values = [
                str(idx + 1),
                str(r.get("name", "")),
                str(r.get("date", "")),
                str(r.get("time", "")),
            ]
            for col_idx, val in enumerate(values):
                is_name = col_idx == 1
                cell = ctk.CTkLabel(
                    self._table_body, text=val,
                    font=ctk.CTkFont(family=FONT_FAMILY, size=12,
                                     weight="bold" if is_name else "normal"),
                    text_color=COLORS["text_primary"] if is_name else COLORS["text_secondary"],
                    fg_color=bg, corner_radius=0, anchor="w",
                )
                cell.grid(row=idx, column=col_idx, padx=14, pady=5, sticky="ew")

    def _read_excel(self, filepath):
        records = []
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=5, values_only=True):
                if row[1] is not None:
                    records.append({
                        "name": str(row[1]),
                        "date": str(row[2]) if row[2] else "",
                        "time": str(row[3]) if row[3] else "",
                    })
            wb.close()
        except Exception:
            pass
        return records

    def _open_file(self):
        if self._current_file and os.path.exists(self._current_file):
            os.startfile(self._current_file)

    def _open_folder(self):
        os.makedirs(ATTENDANCE_FOLDER, exist_ok=True)
        os.startfile(ATTENDANCE_FOLDER)
