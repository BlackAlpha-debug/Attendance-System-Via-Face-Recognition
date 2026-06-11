import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def export_session_to_excel(records, session_start, output_folder):
    os.makedirs(output_folder, exist_ok=True)

    timestamp = session_start.strftime("%Y-%m-%d_%H%M%S")
    filename = f"attendance_{timestamp}.xlsx"
    filepath = os.path.join(output_folder, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1B73E8", end_color="1B73E8", fill_type="solid")
    col_header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    col_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    data_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Attendance Report"
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = center

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Session: {session_start.strftime('%B %d, %Y  %I:%M %p')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10)
    ws["A2"].alignment = center

    columns = ["#", "Student Name", "Date", "Time"]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center
        cell.border = thin_border

    for row_idx, record in enumerate(records, 5):
        ws.cell(row=row_idx, column=1, value=row_idx - 4).font = data_font
        ws.cell(row=row_idx, column=1).alignment = center
        ws.cell(row=row_idx, column=1).border = thin_border

        ws.cell(row=row_idx, column=2, value=record["name"]).font = data_font
        ws.cell(row=row_idx, column=2).border = thin_border

        ws.cell(row=row_idx, column=3, value=record["date"]).font = data_font
        ws.cell(row=row_idx, column=3).alignment = center
        ws.cell(row=row_idx, column=3).border = thin_border

        ws.cell(row=row_idx, column=4, value=record["time"]).font = data_font
        ws.cell(row=row_idx, column=4).alignment = center
        ws.cell(row=row_idx, column=4).border = thin_border

        if row_idx % 2 == 0:
            stripe = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
            for c in range(1, 5):
                ws.cell(row=row_idx, column=c).fill = stripe

    summary_row = len(records) + 6
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    ws.cell(row=summary_row, column=1, value="Total Present:").font = Font(bold=True, size=11)
    ws.cell(row=summary_row, column=3, value=len(records)).font = Font(bold=True, size=11)
    ws.cell(row=summary_row, column=3).alignment = center

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12

    wb.save(filepath)
    return filepath
